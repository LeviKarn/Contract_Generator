from datetime import date, datetime
from pathlib import Path
import os
import platform
import re
import subprocess
import sys
from zipfile import ZipFile

from docxtpl import DocxTemplate
from openpyxl import load_workbook


# Projektordner unabhängig vom aktuellen Speicherort bestimmen
def get_project_dir():
    """
    Determines the folder that contains input/, templates/ and output/.
    In a PyInstaller build this is the folder next to the executable.
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent

    return Path(__file__).resolve().parent.parent


PROJECT_DIR = get_project_dir()

EXCEL_FILE = PROJECT_DIR / "input" / "Contract_Generator.xlsx"
OUTPUT_DIR = PROJECT_DIR / "output"

SHEET_NAME = "Vertragsdaten"

DOCUMENT_TYPES = {
    "rahmenvertrag": {
        "label": "Rahmenvertrag",
        "template": PROJECT_DIR / "templates" / "Rahmenvertrag [aktuelle Version].docx",
        "number_field": "rahmenvertragsnummer",
        "filename_suffix": "Rahmenvertrag",
        "fallback_filename": "Rahmenvertrag_generiert",
    },
    "order_form": {
        "label": "Order Form",
        "template": PROJECT_DIR / "templates" / "Orderform_2026 [aktuelle Version_Juli].docx",
        "number_field": "order_form_nummer",
        "filename_suffix": "Order_Form",
        "fallback_filename": "Order_Form_generiert",
    },
}

TEMPLATE_FILE = DOCUMENT_TYPES["rahmenvertrag"]["template"]


def format_excel_value(value):
    """
    Wandelt Excel-Werte in eine für Word geeignete Darstellung um.
    """
    if value is None:
        return ""

    if isinstance(value, (datetime, date)):
        return value.strftime("%d.%m.%Y")

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def read_contract_data():
    """
    Liest die Eingaben aus Spalte B und ordnet sie anhand der
    technischen Feldnamen aus Spalte D zu.
    """
    workbook = load_workbook(EXCEL_FILE, data_only=True)

    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"Das Tabellenblatt '{SHEET_NAME}' wurde nicht gefunden."
        )

    worksheet = workbook[SHEET_NAME]
    contract_data = {}

    # Zeile 1 enthält die Überschriften
    for row in range(2, worksheet.max_row + 1):
        input_value = worksheet.cell(row=row, column=2).value
        technical_field_name = worksheet.cell(row=row, column=4).value

        if technical_field_name is None:
            continue

        technical_field_name = str(technical_field_name).strip()

        if not technical_field_name:
            continue

        contract_data[technical_field_name] = format_excel_value(input_value)

    return contract_data


def read_field_definitions():
    """
    Reads the field setup from Excel so the app can build its form dynamically.
    """
    workbook = load_workbook(EXCEL_FILE, data_only=True)

    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"Das Tabellenblatt '{SHEET_NAME}' wurde nicht gefunden."
        )

    worksheet = workbook[SHEET_NAME]
    field_definitions = []

    for row in range(2, worksheet.max_row + 1):
        technical_field_name = worksheet.cell(row=row, column=4).value

        if technical_field_name is None:
            continue

        technical_field_name = str(technical_field_name).strip()

        if not technical_field_name:
            continue

        field_definitions.append(
            {
                "label": format_excel_value(worksheet.cell(row=row, column=1).value),
                "default": format_excel_value(worksheet.cell(row=row, column=2).value),
                "example": format_excel_value(worksheet.cell(row=row, column=3).value),
                "technical_name": technical_field_name,
                "hint": format_excel_value(worksheet.cell(row=row, column=5).value),
            }
        )

    return field_definitions


def get_document_type(document_type):
    if document_type not in DOCUMENT_TYPES:
        known_types = ", ".join(DOCUMENT_TYPES)
        raise ValueError(
            f"Unbekannter Dokumenttyp '{document_type}'. Erwartet: {known_types}"
        )

    return DOCUMENT_TYPES[document_type]


def extract_template_placeholders(template_file):
    """
    Reads placeholders from a docx template.
    """
    if not template_file.exists():
        raise FileNotFoundError(
            f"Word-Template nicht gefunden:\n{template_file}"
        )

    xml_parts = []

    with ZipFile(template_file) as archive:
        for name in archive.namelist():
            if name.startswith("word/") and name.endswith(".xml"):
                xml_parts.append(archive.read(name).decode("utf-8", errors="ignore"))

    xml_text = "".join(xml_parts)
    plain_text = re.sub(r"<[^>]+>", "", xml_text)

    return sorted(
        set(re.findall(r"\{\{\s*([a-zA-Z_][a-zA-Z0-9_]*)\s*\}\}", plain_text))
    )


def get_fields_for_document(document_type):
    wanted_placeholders = set(
        extract_template_placeholders(get_document_type(document_type)["template"])
    )
    all_fields = read_field_definitions()

    return [
        field
        for field in all_fields
        if field["technical_name"] in wanted_placeholders
    ]


def sanitize_filename(filename):
    """
    Entfernt Zeichen, die unter Windows oder macOS nicht in
    Dateinamen verwendet werden sollten.
    """
    filename = re.sub(r'[<>:"/\\|?*]', "_", filename)
    filename = filename.strip().strip(".")
    return filename or "Rahmenvertrag_generiert"


def open_folder(folder_path):
    """
    Opens a folder in the user's file browser.
    """
    folder_path = Path(folder_path)

    if platform.system() == "Windows":
        os.startfile(folder_path)
        return

    if platform.system() == "Darwin":
        subprocess.run(["open", str(folder_path)], check=False)
        return

    subprocess.run(["xdg-open", str(folder_path)], check=False)


def clear_output_folder():
    """
    Removes generated Word files from the output folder before a new document is created.
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    for output_file in OUTPUT_DIR.glob("*.docx"):
        output_file.unlink()


def generate_contract(open_output=False, document_type="rahmenvertrag"):
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(
            f"Excel-Datei nicht gefunden:\n{EXCEL_FILE}"
        )

    document_config = get_document_type(document_type)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    contract_data = read_contract_data()
    return render_contract(
        contract_data,
        open_output=open_output,
        document_type=document_type,
    )


def render_contract(
    contract_data,
    open_output=False,
    document_type="rahmenvertrag",
    clear_output=True,
):
    document_config = get_document_type(document_type)
    template_file = document_config["template"]

    if not template_file.exists():
        raise FileNotFoundError(
            f"Word-Template nicht gefunden:\n{template_file}"
        )

    if clear_output:
        clear_output_folder()
    else:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print("\nEingelesene Vertragsdaten:")
    for field_name, value in contract_data.items():
        displayed_value = value if value else "[leer]"
        print(f"  {field_name}: {displayed_value}")

    document = DocxTemplate(template_file)
    document.render(contract_data)

    contract_number = contract_data.get(document_config["number_field"], "")
    output_name = sanitize_filename(
        f"{contract_number}_{document_config['filename_suffix']}"
        if contract_number
        else document_config["fallback_filename"]
    )

    output_file = OUTPUT_DIR / f"{output_name}.docx"
    document.save(output_file)

    print("\nVertrag erfolgreich erstellt:")
    print(output_file)

    if open_output:
        open_folder(OUTPUT_DIR)

    return output_file, contract_data


def main():
    try:
        generate_contract(open_output=True)
    except Exception as error:
        print("\nFEHLER BEI DER VERTRAGSERSTELLUNG")
        print("----------------------------------")
        print(error)
        sys.exit(1)


if __name__ == "__main__":
    main()
